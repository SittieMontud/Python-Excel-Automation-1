#from openpyxl import Workbook
import openpyxl

txtfile = open("D:\Python ex\Automate Excel Tasks\employees.txt", "r")
records = []
txtfile.seek(0)
#Creating a list of lists with the element of the outer list as the row
for record in txtfile.readlines():
	#rstrip -> strips newline in the string
	#split -> splits the string by the delimeter ;
	records.append(record.rstrip("\n").split(";"))

#Create workbook
workbook_path = "D:\\Python ex\\Automate Excel Tasks\\MyCompanyStaff.xlsx"
openpyxl.Workbook().save(workbook_path)

workbookvar = openpyxl.load_workbook(workbook_path)

#Change name of the sheet to Employees
sheetlist = workbookvar.sheetnames
sheetname = sheetlist[0]
sheet = workbookvar[sheetname]
sheet.title = "Employees"


#to transfer data in txt file to workbook
for row in records:
	sheet.append(row)

fontvar = openpyxl.styles.Font(color = '00FF0000', bold = True, italic = False)

#searches which column has the Salary header
colList = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
colSalary = 'A'
for colx in colList:
	if sheet['%s1'%(colx)].value == "Salary":
		colSalary = colx
		break

#highlighting cell rules
for row in range(2,sheet.max_row):
	if int(sheet['%s%s' % (colSalary,row)].value) > 55000:
		sheet['%s%s' % (colSalary,row)].font = fontvar
		
		
#save workbook
workbookvar.save(workbook_path)
